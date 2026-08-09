"""
Reports API — Excel & PDF Export Engine for Orders, Sales, Profit, Inventory, Customers
"""
from django.http import HttpResponse
from rest_framework.views import APIView
from core.permissions import IsAdminUser
from apps.orders.models import Order
from apps.products.models import Product
import openpyxl
from fpdf import FPDF


class OrderExcelExportView(APIView):
    """Export Orders to Excel spreadsheet with proper formatting."""
    permission_classes = [IsAdminUser]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders Report"

        # Headers
        headers = ["Order #", "Customer", "Phone", "Status", "Payment", "Subtotal", "Discount", "Grand Total", "Profit", "Date"]
        ws.append(headers)

        orders = Order.objects.all().order_by('-created_at')[:1000]
        for o in orders:
            ws.append([
                o.order_number, o.customer_name, o.customer_phone, o.status,
                o.payment_method, float(o.subtotal), float(o.discount_amount),
                float(o.grand_total), float(o.estimated_profit),
                o.created_at.strftime("%Y-%m-%d %H:%M")
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="orders_report.xlsx"'
        wb.save(response)
        return response


class OrderPdfInvoiceView(APIView):
    """Generate Order Invoice PDF using fpdf2."""
    permission_classes = [IsAdminUser]

    def get(self, request, order_id):
        try:
            order = Order.objects.prefetch_related('items').get(id=order_id)
        except Order.DoesNotExist:
            return HttpResponse("Order not found", status=404)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, f"INVOICE: {order.order_number}", ln=True, align="C")

        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, f"Store: PriyoShop Enterprise Platform", ln=True)
        pdf.cell(0, 8, f"Customer: {order.customer_name} ({order.customer_phone})", ln=True)
        pdf.cell(0, 8, f"Shipping Address: {order.shipping_address}, {order.shipping_city}", ln=True)
        pdf.cell(0, 8, f"Date: {order.created_at.strftime('%Y-%m-%d')}", ln=True)
        pdf.ln(5)

        # Table header
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(100, 8, "Item", 1)
        pdf.cell(30, 8, "Qty", 1)
        pdf.cell(30, 8, "Price", 1)
        pdf.cell(30, 8, "Total", 1, ln=True)

        pdf.set_font("Helvetica", "", 10)
        for item in order.items.all():
            pdf.cell(100, 8, str(item.product_name)[:40], 1)
            pdf.cell(30, 8, str(item.quantity), 1)
            pdf.cell(30, 8, f"BDT {item.unit_price}", 1)
            pdf.cell(30, 8, f"BDT {item.line_total}", 1, ln=True)

        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, f"Subtotal: BDT {order.subtotal}", ln=True)
        pdf.cell(0, 8, f"Discount: BDT {order.discount_amount}", ln=True)
        pdf.cell(0, 8, f"Shipping: BDT {order.shipping_charge}", ln=True)
        pdf.cell(0, 8, f"Grand Total: BDT {order.grand_total}", ln=True)

        pdf_bytes = pdf.output()
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response['Content-Disposition'] = f'attachment; filename="invoice_{order.order_number}.pdf"'
        return response
